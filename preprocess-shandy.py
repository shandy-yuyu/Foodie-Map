import json
from openpyxl import Workbook

json_name = ""
xlsx_name = ""
for j in range(1, 13):
  json_name = "dataset" + str(j) + ".json"
  xlsx_name = "output" + str(j) + ".xlsx"

  data = []
  with open(json_name) as f:
    data = json.load(f)

  output = []
  for i in range(len(data)):
    output.append([data[i]["name"], data[i]["geometry"]["location"]["lat"], data[i]["geometry"]["location"]["lng"], data[i]["rating"], data[i]["user_ratings_total"], data[i]["vicinity"]])

  wb = Workbook()
  ws = wb.active
  ws['A1'] = 'location-lat'
  ws['B1'] = 'location-lng'
  ws['C1'] = 'name'
  ws['D1'] = 'rating'
  ws['E1'] = 'rating population'
  ws['F1'] = 'address'
  for i in range(len(data)):
    ws.append(output[i])
  wb.save(xlsx_name) 